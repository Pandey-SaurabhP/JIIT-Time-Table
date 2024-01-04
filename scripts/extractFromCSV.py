import openpyxl
import re
import csv
import json

def readCSV(file_path):
    data_dict = {}

    try:
        with open(file_path, 'r', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)

            for row in csv_reader:
                if len(row) == 2:
                    key, value = row
                    data_dict[key.upper()] = value

        return data_dict

    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return None

abb = readCSV('./src/abbreviations.csv')
slots = {0: '09:00-09:50', 1: '10:00-10:50', 2: '11:00-11:50', 3: '12:00-12:50', 4: '13:00-13:50', 5: '14:00-14:50', 6: '15:00-15:50', 7: '16:00-16:50'}
days = {'MON': 'Monday', 'TUES': 'Tuesday', 'WED': 'Wednesday', 'THUR': 'Thursday', 'FRI': 'Friday', 'SAT': 'Saturday'}

# Excel portion
def readExcel(file_path, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        sheet = workbook[sheet_name]
        
        rows = sheet.max_row
        columns = sheet.max_column
        
        data_array = []
        
        for row in range(1, rows + 1):
            row_data = []
            for col in range(1, columns + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            data_array.append(row_data)
        
        return data_array

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None


def extractInfoFromText(course_info, slot, day):

    course_info = course_info.replace(' ', '')
    pattern = re.compile(r'^(T|L|P)(((E|F)[0-9]+)+)\((.*?)\)*-(.+)*/(.+)$')
    
    match = pattern.match(course_info)

    if match:
        course_type = match.group(1)
        batch_info = match.group(2)
        subject_code = match.group(5)
        room_number = match.group(6)
        teacher_name = match.group(7)

        if subject_code in abb:
            subject_code = abb[subject_code]
        if teacher_name in abb:
            teacher_name = abb[teacher_name]
        if slot in slots:
            slot = slots[slot]
        if day in days:
            day = days[day]

        if course_type == 'P':
            course_type = 'Lab'
        elif course_type == 'L':
            course_type = 'Lecture'
        else: course_type = 'Tutorial'


        ret = {
            "Course": course_type,
            "Batch": batch_info,
            "Subject": subject_code,
            "Room": room_number,
            "Teacher": teacher_name,
            "Slot": slot,
            "Day": day
        }

        # print('Result : ', ret)
        return ret
    else:
        return None

finalData = []

def readAndParse():
    result = readExcel('./src/tt.xlsx', 'Sheet2')

    for i in result:
        for j in range(1, len(i)):
            if i[j] is not None and i[j] != 'LUNCH':
                formatted = extractInfoFromText(i[j], j - 1, i[0])
                batches = formatted['Batch']

                for k in range(0, len(batches), 2):
                    batch_data = formatted.copy()
                    batch_data['Batch'] = batches[k: k + 2]
                    finalData.append(batch_data)

    # for i in finalData:
    #     print(i)

def dumpToJSON():
    json_file_path = '../data/data.json'

    with open(json_file_path, 'w') as json_file:
        json.dump(finalData, json_file)

    print(f"Array of dictionaries data saved to {json_file_path}")

readAndParse()
# dumpToJSON()

